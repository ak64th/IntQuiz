# coding=utf-8
import datetime
import math

import excel_tools
import openpyxl
import openpyxl.writer.excel
import os
import redis
import tasks
from api import api
from app import app, db
from archive import ArchiveManager
from auth import auth
from flask import (flash, g, redirect, request, render_template,
                   send_from_directory, url_for, jsonify, make_response)
from flask_peewee.utils import get_object_or_404
from models import *
from models import generate_activity_code
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
from openpyxl.xml.constants import XLSX as XLSX_MIMETYPE
from peewee import create_model_tables, JOIN, fn

auth.setup()
api.setup()

if 'REDIS' in app.config:
    redis_client = redis.StrictRedis.from_url(app.config['REDIS'], decode_responses=True)
else:
    redis_client = redis.StrictRedis(decode_responses=True)

manager = ArchiveManager(redis_client)


@app.route('/')
@auth.login_required
def index():
    return render_template('home.html')


@app.route('/resetPwd', methods=['GET', 'POST'])
@auth.login_required
def reset_password():
    if request.method == 'POST':
        password = request.form.get('password')
        confirm = request.form.get('confirm')
        if password and len(password) < 6:
            flash(u'新密码不能少于六位', 'danger')
        elif password and confirm and password == confirm:
            g.user.set_password(password)
            g.user.save()
            flash(u'修改成功', 'success')
        else:
            flash(u'新密码与确认密码不一致', 'danger')
    return render_template('reset.html')


@app.route('/accounts/create', methods=['POST'])
@auth.admin_required
def create_user():
    data = request.get_json(silent=True)
    user, _ = User.get_or_create(username=data['username'])
    user.set_password(data['password'])
    user.save()
    user_resource = api.registry[User]
    return jsonify(user_resource.serialize_object(user))


@app.route('/accounts/<uid>/changePwd', methods=['POST', 'PUT'])
@auth.admin_required
def change_password(uid):
    data = request.get_json(silent=True)
    user = get_object_or_404(User, (User.id == uid))
    user.username = data['username']
    user.set_password(data['password'])
    user.save()
    user_resource = api.registry[User]
    return jsonify(user_resource.serialize_object(user))


@app.route('/accounts/<uid>/toggle', methods=['POST', 'PUT'])
@auth.admin_required
def toggle_user_state(uid):
    data = request.get_json(silent=True)
    user = get_object_or_404(User, (User.id == uid))
    user.active = data['active']
    user.save()
    user_resource = api.registry[User]
    return jsonify(user_resource.serialize_object(user))


@app.route('/accounts', methods=['GET', 'POST'])
@auth.admin_required
def account_list():
    if request.method == 'POST':
        uid = request.form.get('uid')
        username = request.form.get('username')
        password = request.form.get('password')
        if not (username and password):
            flash(u'缺少字段', 'danger')
        else:
            if uid:
                user = get_object_or_404(User, (User.id == uid))
            else:
                user = User()
            user.username = username
            user.set_password(password)
            if user.save():
                flash(u'操作成功', 'success')
    return render_template('accounts.html')


@app.route('/quizbooks', methods=['GET', 'POST'])
@auth.login_required
def quiz_book_list():
    if request.method == 'POST':
        book_name = request.form.get('name')
        book_file = request.files['file']
        if not book_name:
            flash(u'未输入题库名字', 'warning')
        elif QuizBook.select().where(QuizBook.title == book_name).count():
            flash(u'题库名字重复', 'danger')
        elif book_file:
            handle_sheet_file(book_name, book_file)
        else:
            flash(u'未选择文件', 'warning')
    return render_template('books.html')


def handle_sheet_file(title, f):
    allowed_extensions = ('.xlsx', '.xlsm')
    ext = os.path.splitext(f.filename)[1]
    if ext not in allowed_extensions:
        flash(u'不支持的文件扩展名，只支持{}格式'
              .format(','.join(allowed_extensions)), 'danger')
        return
    wb = openpyxl.load_workbook(f, read_only=True)
    ws = wb.active
    if ws.max_row < 2:
        flash(u'文件不包含题目', 'danger')
        return
    rows = ws.iter_rows("A2:I%s" % ws.max_row)

    validated, error_info = excel_tools.validate_rows(rows)
    missing = error_info.get('missing')
    duplications = error_info.get('duplications')
    seen = error_info.get('seen')
    invalid_type = error_info['invalid_type']
    invalid_correct_option = error_info['invalid_correct_option']
    if missing:
        flash(u'''第{}行缺少字段，必须有题目内容，类型，正确答案和至少一个选项，'''
              .format(','.join(map(str, missing))), 'danger')
    if duplications:
        for k, v in duplications.items():
            flash(u'''第{}行与第{}行内容重复。'''
                  .format(','.join(map(str, v)), seen[k]), 'danger')
    if invalid_type:
        flash(u'''第{}行题目类型错误，必须是“单选”或“多选”，'''
              .format(','.join(map(str, invalid_type))), 'danger')
    if invalid_correct_option:
        flash(u'''
        第{}行正确答案长度和题目类型不符，或是对应答案单元格内容为空。
        注意是不是输入答案时填在下一个单元格了。
        '''.format(','.join(map(str, invalid_correct_option))), 'danger')
    if not any((missing, duplications, invalid_type, invalid_correct_option)):
        database = db.database
        with database.atomic():
            book = QuizBook.create(title=title, user=g.user)
            for row in validated:
                row['book'] = book.id
            Question.insert_many(validated).execute()
        count = Question.select().where(Question.book == book).count()
        flash(u"""保存成功，目前该题库共有{}题""".format(count), 'success')


@app.route('/quizbooks/template')
def quiz_book_template():
    return send_from_directory(
        directory=app.config['MEDIA_PATH'],
        filename=u'template.xlsx',
        as_attachment=True,
        mimetype=XLSX_MIMETYPE)


@app.route('/quizbooks/<book_id>/questions', endpoint='book_question_list')
@app.route('/questions', endpoint='question_list', defaults={'book_id': None})
@auth.login_required
def question_list(book_id):
    if not book_id:
        condition = None if g.user.admin else (QuizBook.user == g.user)
        try:
            book_id = QuizBook.get(condition).id
        except QuizBook.DoesNotExist:
            flash(u'请至少建立一个题库', 'danger')
            return redirect(url_for('quiz_book_list'))
    books = QuizBook.select()
    if not g.user.admin:
        books = books.where(QuizBook.user == g.user)
    return render_template('questions.html', book_id=book_id, books=books)


@app.route('/quizbooks/<book_id>/stats')
def quiz_book_stats(book_id):
    book = get_object_or_404(QuizBook, QuizBook.id == book_id)
    return jsonify(title=book.title, single=book.single_count, multi=book.multi_count)


@app.route('/activities/')
@auth.login_required
def activity_list():
    return render_template('activities.html', front_host=app.config['FRONT_HOST'])


@app.route('/activities/add/', defaults={'pk': None}, endpoint='add_activity', methods=['GET', 'POST'])
@app.route('/activities/edit/<int:pk>/', endpoint='edit_activity', methods=['GET', 'POST'])
@auth.login_required
def activity_detail(pk):
    if pk:
        activity = get_object_or_404(Activity, (Activity.id == pk))
    else:
        condition = None if g.user.admin else (QuizBook.user == g.user)
        try:
            default_book = QuizBook.get(condition)
        except QuizBook.DoesNotExist:
            flash(u'请至少建立一个题库', 'danger')
            return redirect(url_for('activity_list'))

        activity = Activity(type=0,
                            book=default_book,
                            start_at=datetime.datetime.now(),
                            show_answer=True,
                            code=generate_activity_code())

    if request.method == 'POST':
        name = request.form.get('name')
        welcome = request.form.get('welcome')
        _type = request.form.get('type', type=int)
        book = request.form.get('book', type=int)
        chances = request.form.get('chances', 0, type=int)
        time_limit = request.form.get('time_limit', 0, type=int)
        single = request.form.get('single', 0, type=int)
        multi = request.form.get('multi', 0, type=int)
        single_points = request.form.get('single_points', 0, type=int)
        multi_points = request.form.get('multi_points', 0, type=int)
        show_answer = request.form.get('show_answer', False, type=bool)
        datetimerange = request.form.get('datetimerange')
        info_field_1 = request.form.get('info_field_1')
        info_field_2 = request.form.get('info_field_2')
        info_field_3 = request.form.get('info_field_3')
        welcome_img = request.files['welcome_img']

        valid = True
        if not all((name, book, chances, datetimerange)):
            flash(u'字段缺失', 'danger')
            valid = False
        if chances and chances < 1:
            flash(u'答题次数不能小于1', 'danger')
            valid = False
        if _type == Activity.ORDINARY:
            activity.book = get_object_or_404(QuizBook, (QuizBook.id == book))
            if single + multi < 1:
                flash(u'普通模式下，题目总数不能小于1', 'danger')
                valid = False
            single_count = activity.book.questions.where(Question.type == Question.SINGLE).count()
            multi_count = activity.book.questions.where(Question.type == Question.MULTI).count()
            if single_count < single or multi_count < multi:
                flash(u'设置的题目数量多于题库中题目数量，目前题库中共有单选{}题，多选{}题'
                      .format(single_count, multi_count), 'danger')
                valid = False
        if any(map(lambda x: x < 0, [time_limit, single, multi, single_points, multi_points])):
            flash(u'时间限制，题目数量和分值不能为负', 'danger')
            valid = False
        if single_points + multi_points < 1:
            flash(u'单选和多选分值不能同时为0', 'danger')
            valid = False

        if welcome_img:
            try:
                activity.welcome_img = tasks.save_welcome_img(welcome_img, activity)
            except tasks.UploadNotAllowed as e:
                flash(e.message, 'danger')
                valid = False

        def datetime_format(date_string):
            return datetime.datetime.strptime(date_string.strip(), u'%Y-%m-%d %H:%M:%S')

        if datetimerange:
            try:
                start_at, end_at = map(datetime_format, datetimerange.split(' - '))
            except ValueError:
                flash(u'日期时间字符串格式错误', 'danger')
                valid = False

        # 把非空白的字段放到数组最前，然后接上None
        info_fields = list(filter(lambda x: x.strip() if x else x,
                                  (info_field_1, info_field_2, info_field_3))) + [None] * 3

        if valid:
            activity.name = name
            activity.welcome = welcome
            activity.type = _type
            activity.chances = chances
            activity.time_limit = time_limit
            activity.single = single
            activity.multi = multi
            activity.single_points = single_points
            activity.multi_points = multi_points
            activity.show_answer = show_answer
            activity.start_at = start_at
            activity.end_at = end_at
            activity.info_field_1 = info_fields[0]
            activity.info_field_2 = info_fields[1]
            activity.info_field_3 = info_fields[2]
            activity.user = g.user
            if activity.save():
                tasks.generate_json_files_for_activity(activity)
                flash(u'保存成功', 'success')
                return redirect(url_for('activity_list'))
    books = QuizBook.select()
    return render_template('activity.html', activity=activity, books=books)


@app.route('/activities/<int:pk>/delete_img/')
@auth.login_required
def delete_welcome_image(pk):
    activity = get_object_or_404(Activity, (Activity.id == pk))
    activity.welcome_img = None
    if activity.save():
        tasks.generate_json_files_for_activity(activity)
        flash(u'删除图片成功', 'success')
    return redirect(url_for('edit_activity', pk=pk))


@app.route('/activities/publish/<int:pk>/')
def publish_activity(pk):
    activity = get_object_or_404(Activity, (Activity.id == pk))
    tasks.upload_files_for_activity(activity)
    flash(u'已经将活动{}的题目分发到前台节点'.format(activity.name), 'success')
    return redirect(url_for('activity_list'))


@app.route('/welcome_image/<filename>')
def welcome_image(filename):
    path = os.path.join(app.config['MEDIA_PATH'], 'welcome_image')
    resp = send_from_directory(path, filename)
    resp.headers.add('Last-Modified', datetime.datetime.now())
    resp.headers.add('Cache-Control', 'no-store, no-cache, must-revalidate, post-check=0, pre-check=0')
    resp.headers.add('Pragma', 'no-cache')
    return resp


@app.route('/stats/')
@auth.login_required
def activity_stats_list():
    query = (Activity
             .select(Activity.id,
                     Activity.name,
                     Activity.start_at,
                     Activity.end_at,
                     Activity.user,
                     fn.Count(Archive.run_id).alias('run_count'),
                     fn.Count(fn.DISTINCT(Archive.uid)).alias('user_count'))
             .join(Archive, JOIN.LEFT_OUTER)
             .group_by(Activity.id, Activity.name)
             .order_by(Activity.created.desc())
             .naive())
    if not g.user.admin:
        query = query.where(Activity.user == g.user)
    return render_template('stats.html', activities=query)


@app.route('/stats/<int:activity_id>/', methods=['GET', 'POST'])
@auth.login_required
def activity_stats_detail(activity_id):
    activity = Activity.get(Activity.id == activity_id)
    if request.method == 'POST':
        manager.archive(activity)
    page = request.args.get('page', 1, type=int)
    paginate_by = request.args.get('limit', 20, type=int)
    query = activity.archives.order_by(Archive.id)
    pages = int(math.ceil(float(query.count()) / paginate_by))
    runs = query.paginate(page=page, paginate_by=paginate_by)
    return render_template('stats_details.html', activity_id=activity_id, runs=runs,
                           cur_page=page, paginate_by=paginate_by, pages=pages)


def remove_illgeal_characters(s):
    if not s:
        return u''
    # never 502
    # noinspection PyBroadException
    try:
        try:
            return ILLEGAL_CHARACTERS_RE.sub(r'', s)
        except UnicodeEncodeError:
            return ILLEGAL_CHARACTERS_RE.sub(r'', s.decode('utf-8'))
    except Exception:
        return u''


@app.route('/stats/<int:activity_id>/workbook')
@auth.login_required
def activity_stats_workbook(activity_id):
    activity = Activity.get(Activity.id == activity_id)
    title = u'-'.join([u'活动详情', activity.name])
    wb = openpyxl.Workbook(guess_types=True)
    ws = wb.active
    ws.title = title
    ws.append([u'名次', u'用户字段1', u'用户字段2', u'用户字段3', u'分数', u'提交成绩时间'])
    for i, archive in enumerate(activity.archives.order_by(Archive.score.desc()), 1):
        ws.append([i,
                   remove_illgeal_characters(archive.info_field_1),
                   remove_illgeal_characters(archive.info_field_2),
                   remove_illgeal_characters(archive.info_field_3),
                   archive.score,
                   archive.end])
    virtual_workbook = openpyxl.writer.excel.save_virtual_workbook(wb)
    response = make_response(virtual_workbook)
    response.mimetype = XLSX_MIMETYPE
    response.headers.add(u'Content-Disposition',
                         u'attachment; filename=activity_%s.xlsx' % activity.id)
    return response


def init_db():
    to_create = User, QuizBook, Question, Activity, UserInfo, Run, FinalScore, Archive
    create_model_tables(models=to_create, fail_silently=True)
    # 创建测试用户
    user, created = User.get_or_create(username='admin', admin=True)
    if created:
        user.set_password('123456')
        user.save()


if __name__ == '__main__':
    import logging
    import logging.handlers
    logger = logging.getLogger('peewee')
    logger.setLevel(logging.DEBUG)
    stream_handler = logging.StreamHandler()
    stream_handler.setFormatter(logging.Formatter("%(asctime)s - %(name)s - %(levelname)s - %(message)s"))
    logger.addHandler(stream_handler)

    init_db()
    app.run(host='0.0.0.0', debug=True, threaded=True)